import logging
logger = logging.getLogger(__name__)
"""
Blueprint de Préstamos con Sistema de Notificaciones
====================================================
Este archivo debe reemplazar: blueprints/prestamos.py

Cambios principales:
- Integración con NotificationService
- Notificaciones en: crear, aprobar, rechazar, devolver
"""

from flask import Blueprint, render_template, request, redirect, url_for, flash, session, send_file, after_this_request, jsonify, current_app
from utils.permissions import can_access
from io import BytesIO
from datetime import datetime, timedelta
from decimal import Decimal
import pandas as pd
import tempfile
import os
from werkzeug.utils import secure_filename
from werkzeug.routing import BuildError
from jinja2 import TemplateNotFound

# Helpers seguros (sanitización / formato) con fallback defensivo
try:
    from utils.helpers import sanitizar_log_text, format_currency, format_date
except Exception:
    def sanitizar_log_text(value, max_len=500):
        if value is None:
            return ''
        try:
            s = "{0}".format(value)
        except Exception:
            return '[texto-protegido]'
        s = s.replace('\r', '\\r').replace('\n', '\\n').replace('\t', '\\t')
        # Remover otros controles ASCII (<32) excepto espacio
        s = ''.join(ch for ch in s if (ord(ch) >= 32) or ch == ' ')
        if max_len and len(s) > max_len:
            s = s[:max_len] + '...'
        return s

    def format_currency(value):
        try:
            if value is None:
                return "$0"
            return f"${float(value):,.0f}".replace(",", ".")
        except Exception:
            return "$0"

    def format_date(date_value, format_str='%d/%m/%Y'):
        try:
            if not date_value:
                return ""
            if isinstance(date_value, str):
                return date_value
            return date_value.strftime(format_str)
        except Exception:
            return "{0}".format(date_value) if date_value else ""

# Import defensivo para dependencias opcionales
try:
    from weasyprint import HTML
    HAS_WEASYPRINT = True
except ImportError:
    HAS_WEASYPRINT = False

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    pd = None
    HAS_PANDAS = False

from database import get_database_connection

# ====================================
# IMPORTAR SERVICIO DE NOTIFICACIONES
# ====================================
try:
    from services.notification_service import NotificationService
    NOTIFICACIONES_ACTIVAS = True
    logger.info("✅ Servicio de notificaciones cargado para préstamos")
except ImportError:
    NOTIFICACIONES_ACTIVAS = False
    logger.info("⚠️ Servicio de notificaciones no disponible para préstamos")

prestamos_bp = Blueprint('prestamos', __name__)

# =========================
# Render seguro de templates + reparación UTF-8 (patrón del sistema)
# =========================

def _project_root():
    """Determina la raíz del proyecto para ubicar /templates/ de forma robusta."""
    here = os.path.abspath(os.path.dirname(__file__))
    if os.path.basename(here).lower() == 'blueprints':
        return os.path.abspath(os.path.join(here, '..'))
    return here

def _template_path(template_name: str) -> str:
    return os.path.join(_project_root(), 'templates', template_name)

def _ensure_template_utf8(template_name: str):
    """Intenta convertir una plantilla a UTF-8 (útil en Windows/cp1252)."""
    path = _template_path(template_name)
    if not os.path.exists(path):
        return False, "template_not_found"

    # 1) Si ya es UTF-8, no hacer nada
    try:
        with open(path, 'r', encoding='utf-8') as f:
            f.read()
        return True, "already_utf8"
    except UnicodeDecodeError:
        pass

    # 2) Intentar encodings típicos
    for enc in ('cp1252', 'latin-1'):
        try:
            with open(path, 'r', encoding=enc, errors='strict') as f:
                content = f.read()
            with open(path, 'w', encoding='utf-8', newline='\n') as f:
                f.write(content)
            return True, f"converted_from_{enc}"
        except Exception:
            continue

    # 3) Último intento: reemplazo de caracteres inválidos
    try:
        with open(path, 'r', encoding='cp1252', errors='replace') as f:
            content = f.read()
        with open(path, 'w', encoding='utf-8', newline='\n') as f:
            f.write(content)
        return True, "converted_with_replacement"
    except Exception as e:
        return False, "convert_failed"

def safe_url_for(endpoint: str, **values):
    """url_for seguro para templates de préstamos.
    Evita que un endpoint faltante (BuildError) rompa el render.
    Retorna '#' si no se puede construir la URL.
    """
    try:
        if not endpoint:
            return '#'
        return url_for(endpoint, **values)
    except BuildError:
        logger.warning(
            "url_for(BuildError) en template prestamos: endpoint=%s",
            sanitizar_log_text(endpoint)
        )
        return '#'
    except Exception as e:
        logger.warning(
            "url_for(error) en template prestamos: endpoint=%s",
            sanitizar_log_text(endpoint)
        )
        return '#'

def safe_render_template(template_name: str, **context):
    """
    Renderiza una plantilla de forma segura:
    - Si hay UnicodeDecodeError, intenta convertir la plantilla a UTF-8 y reintenta.
    - Evita 500 por TemplateNotFound con un fallback mínimo (UTF-8).
    - Inyecta helpers comunes (format_currency/format_date) para evitar UndefinedError en templates.
    """
    try:
        context.setdefault('format_currency', format_currency)
        context.setdefault('format_date', format_date)
        context.setdefault('url_for', safe_url_for)
        return render_template(template_name, **context)
    except UnicodeDecodeError:
        ok, info = _ensure_template_utf8(template_name)
        if ok:
            try:
                return render_template(template_name, **context)
            except Exception as e:
                logger.error(
                    "Error renderizando template '%s' tras reparación UTF-8: [error]",
                    sanitizar_log_text(template_name)
                )
        else:
            logger.error(
                "No se pudo reparar UTF-8 para template '%s': %s",
                sanitizar_log_text(template_name),
                sanitizar_log_text(info)
            )
        return (
            """<!doctype html><html lang="es"><head><meta charset="utf-8"><title>Error</title></head>
            <body><h3>Error interno</h3><p>No fue posible renderizar la vista solicitada.</p></body></html>""",
            500
        )
    except TemplateNotFound:
        logger.error("Template no encontrada: %s", sanitizar_log_text(template_name))
        return (
            """<!doctype html><html lang="es"><head><meta charset="utf-8"><title>Error</title></head>
            <body><h3>Error interno</h3><p>Vista no disponible.</p></body></html>""",
            500
        )
    except Exception as e:
        logger.error(
            "Error renderizando template '%s': [error]",
            sanitizar_log_text(template_name)
        )
        return (
            """<!doctype html><html lang="es"><head><meta charset="utf-8"><title>Error</title></head>
            <body><h3>Error interno</h3><p>No fue posible procesar la solicitud.</p></body></html>""",
            500
        )

# =========================
# Helpers de sesión / permisos
# =========================
def _require_login():
    return 'usuario_id' in session

def _has_role(*roles):
    rol = (session.get('rol', '') or '').strip().lower()
    return rol in [r.lower() for r in roles]

def _redirect_login():
    """Redirección consistente al login del sistema.

    Seguridad:
    - Evita BuildError si el nombre del endpoint cambia.
    - No construye redirecciones externas (solo same-origin).
    """
    next_url = request.url if request else ''
    for ep in ('auth.login', 'auth_bp.login'):
        try:
            return redirect(url_for(ep, next=next_url))
        except Exception:
            continue
    return redirect('/auth/login')

# =========================
# Helpers de imágenes
# =========================
IMG_COLS = ["RutaImagen", "ImagenURL", "ImagenUrl", "Imagen", "FotoURL", "FotoUrl", "Foto"]

def _detect_image_column(cur):
    """Detecta la primera columna de imagen disponible en ElementosPublicitarios."""
    cur.execute("SELECT TOP 1 * FROM dbo.ElementosPublicitarios")
    col_names = [d[0] for d in cur.description]
    for c in IMG_COLS:
        if c in col_names:
            return c
    return None

def _normalize_image_url(path_value: str) -> str:
    """Normaliza valores de imagen a una URL servible por Flask static."""
    if not path_value:
        return ""
    if isinstance(path_value, str) and path_value.startswith('http'):
        return path_value
    if isinstance(path_value, str) and path_value.startswith('static/'):
        rel = path_value.replace('static/', '')
        return url_for('static', filename=rel)
    if isinstance(path_value, str):
        return url_for('static', filename=path_value)
    return ""

def allowed_file(filename):
    ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# ==============================
# Funciones para notificaciones
# ==============================
def _obtener_info_prestamo_completa(prestamo_id):
    """Obtiene información completa del préstamo para notificaciones"""
    conn = cur = None
    try:
        conn = get_database_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT 
                pe.PrestamoId,
                el.NombreElemento,
                pe.CantidadPrestada,
                u.NombreUsuario,
                u.CorreoElectronico,
                o.NombreOficina,
                pe.Evento,
                pe.FechaDevolucionPrevista,
                pe.Estado
            FROM dbo.PrestamosElementos pe
            INNER JOIN dbo.ElementosPublicitarios el ON pe.ElementoId = el.ElementoId
            INNER JOIN dbo.Usuarios u ON pe.UsuarioSolicitanteId = u.UsuarioId
            INNER JOIN dbo.Oficinas o ON pe.OficinaId = o.OficinaId
            WHERE pe.PrestamoId = ? AND pe.Activo = 1
        """, (prestamo_id,))
        row = cur.fetchone()
        if row:
            return {
                'id': row[0],
                'material': row[1],
                'cantidad': row[2],
                'solicitante_nombre': row[3],
                'email_solicitante': row[4],
                'oficina_nombre': row[5],
                'evento': row[6],
                'fecha_prevista': str(row[7]) if row[7] else 'N/A',
                'estado': row[8]
            }
        return None
    except Exception as e:
        logger.error("Error obteniendo info préstamo: [error]")
        return None
    finally:
        if cur: cur.close()
        if conn: conn.close()

# =========================
# Consultas de base de datos
# =========================
def _fetch_estados_distintos():
    conn = cur = None
    try:
        conn = get_database_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT DISTINCT Estado
            FROM dbo.PrestamosElementos
            WHERE Activo = 1
            ORDER BY Estado
        """)
        return [row[0] for row in cur.fetchall() if row and row[0]]
    except Exception as e:
        logger.error("Error leyendo estados: [error]")
        return []
    finally:
        try:
            if cur: cur.close()
            if conn: conn.close()
        except:
            pass

def _fetch_prestamos(estado=None, oficina_id=None):
    """Lista préstamos con filtro opcional por estado y oficina."""
    conn = cur = None
    rows_out = []
    try:
        conn = get_database_connection()
        cur = conn.cursor()

        sql = """
            SELECT 
                pe.PrestamoId               AS Id,
                pe.ElementoId               AS ElementoId,
                el.NombreElemento           AS Material,
                el.ValorUnitario            AS ValorUnitario,
                pe.CantidadPrestada         AS Cantidad,
                u.NombreUsuario             AS SolicitanteNombre,
                o.NombreOficina             AS OficinaNombre,
                pe.FechaPrestamo            AS Fecha,
                pe.FechaDevolucionPrevista  AS FechaPrevista,
                pe.Estado                   AS Estado,
                pe.Observaciones            AS Observaciones,
                pe.UsuarioAprobador         AS UsuarioAprobador,
                pe.FechaAprobacion          AS FechaAprobacion,
                pe.UsuarioRechazador        AS UsuarioRechazador,
                pe.FechaRechazo             AS FechaRechazo,
                pe.UsuarioDevolucion        AS UsuarioDevolucion,
                pe.FechaDevolucionReal      AS FechaDevolucionReal,
                pe.OficinaId               AS OficinaId
            FROM dbo.PrestamosElementos pe
            INNER JOIN dbo.ElementosPublicitarios el
                ON el.ElementoId = pe.ElementoId
            INNER JOIN dbo.Usuarios u
                ON u.UsuarioId = pe.UsuarioSolicitanteId
            INNER JOIN dbo.Oficinas o
                ON o.OficinaId = pe.OficinaId
            WHERE pe.Activo = 1
        """
        params = []
        
        if oficina_id:
            sql += " AND pe.OficinaId = ?"
            params.append(oficina_id)
        
        if estado and estado.strip():
            sql += " AND pe.Estado = ?"
            params.append(estado.strip())

        sql += " ORDER BY pe.FechaPrestamo DESC"

        cur.execute(sql, params)
        rows = cur.fetchall()

        for r in rows:
            id_ = r[0]
            valor_unit = r[3] or 0
            cant = r[4] or 0
            subtotal = Decimal(valor_unit) * Decimal(cant)
            rows_out.append({
                'id': id_,
                'elemento_id': r[1],
                'material': r[2],
                'valor_unitario': Decimal(valor_unit),
                'cantidad': int(cant),
                'subtotal': subtotal,
                'solicitante_nombre': r[5] or 'N/A',
                'oficina_nombre': r[6] or 'N/A',
                'oficina_id': int(r[17]) if len(r) > 17 and r[17] is not None else None,
                'fecha': r[7],
                'fecha_prevista': r[8],
                'estado': r[9] or '',
                'observaciones': r[10] or '',
                'usuario_aprobador': r[11] or '',
                'fecha_aprobacion': r[12],
                'usuario_rechazador': r[13] or '',
                'fecha_rechazo': r[14],
                'usuario_devolucion': r[15] or '',
                'fecha_devolucion_real': r[16]
            })
    except Exception as e:
        logger.error("Error leyendo préstamos: [error]")
        flash("Error interno al leer préstamos. Intenta de nuevo o contacta al administrador.", "danger")
    finally:
        try:
            if cur: cur.close()
            if conn: conn.close()
        except:
            pass
    return rows_out

def _fetch_detalle(prestamo_id: int):
    conn = cur = None
    try:
        conn = get_database_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT 
                pe.PrestamoId,
                pe.ElementoId,
                el.NombreElemento,
                el.ValorUnitario,
                pe.CantidadPrestada,
                u.NombreUsuario,
                o.NombreOficina,
                pe.FechaPrestamo,
                pe.FechaDevolucionPrevista,
                pe.FechaDevolucionReal,
                pe.Estado,
                pe.Observaciones,
                pe.UsuarioAprobador,
                pe.FechaAprobacion,
                pe.UsuarioRechazador,
                pe.FechaRechazo,
                pe.UsuarioDevolucion,
                pe.MotivoRechazo,
                pe.ObservacionesAprobacion
            FROM dbo.PrestamosElementos pe
            INNER JOIN dbo.ElementosPublicitarios el
                ON el.ElementoId = pe.ElementoId
            INNER JOIN dbo.Usuarios u
                ON u.UsuarioId = pe.UsuarioSolicitanteId
            INNER JOIN dbo.Oficinas o
                ON o.OficinaId = pe.OficinaId
            WHERE pe.PrestamoId = ? AND pe.Activo = 1
        """, (prestamo_id,))
        row = cur.fetchone()
        if not row:
            return None
        valor_unit = Decimal(row[3] or 0)
        cant = int(row[4] or 0)
        return {
            'id': row[0],
            'elemento_id': row[1],
            'material': row[2],
            'valor_unitario': valor_unit,
            'cantidad': cant,
            'subtotal': valor_unit * cant,
            'solicitante_nombre': row[5] or 'N/A',
            'oficina_nombre': row[6] or 'N/A',
            'fecha': row[7],
            'fecha_prevista': row[8],
            'fecha_real': row[9],
            'estado': row[10] or '',
            'observaciones': row[11] or '',
            'usuario_aprobador': row[12] or '',
            'fecha_aprobacion': row[13],
            'usuario_rechazador': row[14] or '',
            'fecha_rechazo': row[15],
            'usuario_devolucion': row[16] or '',
            'motivo_rechazo': row[17] or '',
            'observaciones_aprobacion': row[18] or ''
        }
    except Exception as e:
        logger.error("Error leyendo detalle: [error]")
        return None
    finally:
        try:
            if cur: cur.close()
            if conn: conn.close()
        except:
            pass

# =========================
# Filtros para exportaciones
# =========================
def filtrar_por_oficina_usuario_prestamos(prestamos, campo_oficina='oficina_id'):
    """Filtra préstamos por oficina del usuario actual"""
    from utils.permissions import user_can_view_all
    
    if user_can_view_all():
        return prestamos
    
    oficina_usuario = session.get('oficina_id')
    if not oficina_usuario:
        return []
    
    return [p for p in prestamos if p.get(campo_oficina) == oficina_usuario]

def _fetch_oficinas():
    """Devuelve listado de oficinas para filtros (id, nombre).

    Nota: No asume columnas adicionales; se limita a OficinaId y NombreOficina.
    """
    conn = cur = None
    try:
        conn = get_database_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT OficinaId, NombreOficina
            FROM dbo.Oficinas
            ORDER BY NombreOficina
        """)
        return [{'id': int(r[0]), 'nombre': str(r[1])} for r in (cur.fetchall() or []) if r and r[0] is not None]
    except Exception as e:
        logger.error("Error leyendo oficinas: [error]")
        return []
    finally:
        try:
            if cur: cur.close()
            if conn: conn.close()
        except:
            pass

def _parse_ymd(s: str):
    """Parsea fecha YYYY-MM-DD a datetime (00:00) o None si inválida."""
    try:
        if not s:
            return None
        return datetime.strptime(s, '%Y-%m-%d')
    except Exception:
        return None

def _apply_extra_filters(prestamos, filtro_material='', filtro_solicitante='', fecha_inicio='', fecha_fin=''):
    """Aplica filtros adicionales sobre la lista de préstamos (en memoria)."""
    mat = (filtro_material or '').strip().lower()
    sol = (filtro_solicitante or '').strip().lower()

    dt_ini = _parse_ymd((fecha_inicio or '').strip())
    dt_fin = _parse_ymd((fecha_fin or '').strip())
    if dt_fin:
        # incluir todo el día fin
        dt_fin = dt_fin + timedelta(days=1)

    out = []
    for p in prestamos or []:
        try:
            if mat:
                material = (p.get('material') or '').lower()
                if mat not in material:
                    continue

            if sol:
                solicitante = (p.get('solicitante_nombre') or '').lower()
                if sol not in solicitante:
                    continue

            if dt_ini or dt_fin:
                f = p.get('fecha')
                if not isinstance(f, datetime):
                    # intentar parsear si viene como string
                    try:
                        f = datetime.fromisoformat(f if isinstance(f, str) else "{0}".format(f))
                    except Exception:
                        f = None
                if dt_ini and (not f or f < dt_ini):
                    continue
                if dt_fin and (not f or f >= dt_fin):
                    continue

            out.append(p)
        except Exception:
            # si un registro viene con datos inesperados, se omite sin romper la vista
            continue
    return out


# =========================
# Rutas principales
# =========================
@prestamos_bp.route('/')
def listar_prestamos():
    """Listar préstamos (con filtros).

    Filtros soportados (GET):
    - estado
    - oficina (solo si el usuario puede ver todas)
    - material (búsqueda parcial)
    - solicitante (búsqueda parcial)
    - fecha_inicio (YYYY-MM-DD)
    - fecha_fin (YYYY-MM-DD)
    """
    if not _require_login():
        return _redirect_login()

    filtro_estado = (request.args.get('estado') or '').strip()
    filtro_oficina = (request.args.get('oficina') or 'todas').strip() or 'todas'
    filtro_material = (request.args.get('material') or '').strip()
    filtro_solicitante = (request.args.get('solicitante') or '').strip()
    filtro_fecha_inicio = (request.args.get('fecha_inicio') or '').strip()
    filtro_fecha_fin = (request.args.get('fecha_fin') or '').strip()

    from utils.permissions import user_can_view_all
    can_all = user_can_view_all()

    oficinas = []
    oficina_id = None

    if can_all:
        oficinas = _fetch_oficinas()
        if filtro_oficina and filtro_oficina != 'todas':
            try:
                oficina_id = int(filtro_oficina)
            except Exception:
                oficina_id = None
                filtro_oficina = 'todas'
    else:
        oficina_id = session.get('oficina_id')
        # Para consistencia, reflejar la oficina del usuario en el filtro (aunque el UI no la muestre)
        try:
            filtro_oficina = str(int(oficina_id)) if oficina_id is not None else ''
        except Exception:
            filtro_oficina = ''

    estado_sql = filtro_estado or None
    prestamos = _fetch_prestamos(estado_sql, oficina_id)

    # Filtros adicionales en memoria
    prestamos = _apply_extra_filters(
        prestamos,
        filtro_material=filtro_material,
        filtro_solicitante=filtro_solicitante,
        fecha_inicio=filtro_fecha_inicio,
        fecha_fin=filtro_fecha_fin
    )

    estados = _fetch_estados_distintos()

    return safe_render_template(
        'prestamos/listar.html',
        prestamos=prestamos,
        oficinas=oficinas,
        filtro_estado=filtro_estado,
        filtro_oficina=filtro_oficina,
        filtro_material=filtro_material,
        filtro_solicitante=filtro_solicitante,
        filtro_fecha_inicio=filtro_fecha_inicio,
        filtro_fecha_fin=filtro_fecha_fin,
        estados=estados
    )

@prestamos_bp.route('/crear', methods=['GET', 'POST'])
def crear_prestamo():
    """Crear nuevo préstamo"""
    if not _require_login():
        return _redirect_login()
    
    if not can_access('prestamos', 'create'):
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': 'No tienes permisos para crear préstamos'}), 403
        flash('No tienes permisos para crear préstamos', 'danger')
        return redirect('/prestamos')

    if request.method == 'POST':
        solicitante_id = int(session.get('usuario_id', 0))
        oficina_id = int(session.get('oficina_id', 0))

        elemento_id = request.form.get('elemento_id')
        cantidad = request.form.get('cantidad') or '0'
        fecha_prevista = request.form.get('fecha_prevista')
        evento = (request.form.get('evento') or '').strip()
        observaciones = (request.form.get('observaciones') or '').strip()

        # Validaciones
        if not elemento_id:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'message': 'Debes seleccionar un elemento'})
            flash('Debes seleccionar un elemento', 'warning')
            return redirect('/prestamos/crear')
        if int(cantidad) <= 0:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'message': 'La cantidad debe ser mayor a 0'})
            flash('La cantidad debe ser mayor a 0', 'warning')
            return redirect('/prestamos/crear')
        if not fecha_prevista:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'message': 'La fecha de devolución prevista es obligatoria'})
            flash('La fecha de devolución prevista es obligatoria', 'warning')
            return redirect('/prestamos/crear')
        if not evento:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'message': 'El evento/motivo del préstamo es obligatorio'})
            flash('El evento/motivo del préstamo es obligatorio', 'warning')
            return redirect('/prestamos/crear')
        if not observaciones:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'message': 'Las observaciones son obligatorias'})
            flash('Las observaciones son obligatorias', 'warning')
            return redirect('/prestamos/crear')
        if not solicitante_id or not oficina_id:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'message': 'No se encontraron datos de sesión para solicitante/oficina'})
            flash('No se encontraron datos de sesión para solicitante/oficina', 'danger')
            return redirect('/prestamos/crear')

        conn = cur = None
        try:
            conn = get_database_connection()
            cur = conn.cursor()
            
            # Valida stock
            cur.execute("""
                SELECT CantidadDisponible, NombreElemento
                FROM dbo.ElementosPublicitarios WITH (UPDLOCK, ROWLOCK)
                WHERE ElementoId = ? AND Activo = 1
            """, (int(elemento_id),))
            row = cur.fetchone()
            if not row:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return jsonify({'success': False, 'message': 'Elemento no encontrado o inactivo'})
                flash('Elemento no encontrado o inactivo', 'danger')
                return redirect('/prestamos/crear')

            disponible = int(row[0] or 0)
            nombre_elemento = row[1]
            cantidad_int = int(cantidad)
            
            if cantidad_int > disponible:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return jsonify({'success': False, 'message': f'Stock insuficiente. Disponible: {disponible}'})
                flash(f'Stock insuficiente. Disponible: {disponible}', 'danger')
                return redirect('/prestamos/crear')

            usuario_prestador = session.get('usuario_nombre', 'Sistema')

            
            cur.execute("""
                INSERT INTO dbo.PrestamosElementos
                    (ElementoId, UsuarioSolicitanteId, OficinaId, CantidadPrestada, 
                     FechaPrestamo, FechaDevolucionPrevista, Estado, Evento, Observaciones, 
                     UsuarioPrestador, Activo)
                OUTPUT INSERTED.PrestamoId
                VALUES (?, ?, ?, ?, GETDATE(), ?, 'PENDIENTE', ?, ?, ?, 1)
            """, (
                int(elemento_id), solicitante_id, oficina_id, cantidad_int,
                fecha_prevista, evento, observaciones, usuario_prestador
            ))

            prestamo_id = cur.fetchone()[0]
            logger.info("✅ Préstamo creado con ID: %s", sanitizar_log_text(prestamo_id))
            # Descontar stock
            cur.execute("""
                UPDATE dbo.ElementosPublicitarios
                SET CantidadDisponible = CantidadDisponible - ?
                WHERE ElementoId = ? AND Activo = 1
            """, (cantidad_int, int(elemento_id)))

            conn.commit()
            
            # ====== NOTIFICACIÓN: Préstamo creado ======
            if NOTIFICACIONES_ACTIVAS:
                try:
                    prestamo_info = _obtener_info_prestamo_completa(prestamo_id)
                    if prestamo_info:
                        NotificationService.notificar_prestamo_creado(prestamo_info)
                        logger.info("📧 Notificación enviada: Nuevo préstamo #%s", sanitizar_log_text(prestamo_id))
                except Exception as e:
                    logger.error("Error enviando notificación de préstamo creado: [error]")
            # =============================================
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({
                    'success': True,
                    'message': f'✅ Préstamo registrado exitosamente: {nombre_elemento} (Cantidad: {cantidad})',
                    'prestamo_id': prestamo_id,
                    'elemento_nombre': nombre_elemento,
                    'cantidad': cantidad,
                    'redirect': '/prestamos'
                })
                
            flash(f'✅ Préstamo de "{nombre_elemento}" registrado correctamente para el evento: {evento}', 'success')
            return redirect('/prestamos')
            
        except Exception as e:
            try:
                if conn: conn.rollback()
            except:
                pass
            
            logger.info("❌ Error en crear_prestamo: [error]")
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'message': 'Error interno al crear el préstamo. Intenta nuevamente.'}), 500
            flash('Error interno al crear el préstamo. Intenta nuevamente.', 'danger')
            return redirect('/prestamos/crear')
        finally:
            try:
                if cur: cur.close()
                if conn: conn.close()
            except:
                pass

    # GET: Mostrar formulario
    solicitante_id = session.get('usuario_id', 0)
    solicitante_nombre = session.get('usuario_nombre', '—')
    oficina_id = session.get('oficina_id', 0)
    oficina_nombre = session.get('oficina_nombre', '—')

    fecha_minima = datetime.now().strftime('%Y-%m-%d')

    elementos = []
    conn = cur = None
    try:
        conn = get_database_connection()
        cur = conn.cursor()
        
        img_col = _detect_image_column(cur)

        if img_col:
            cur.execute(f"""
                SELECT ElementoId, NombreElemento, ValorUnitario, CantidadDisponible, {img_col}
                FROM dbo.ElementosPublicitarios
                WHERE Activo = 1 AND CantidadDisponible > 0
                ORDER BY NombreElemento
            """)
            for (eid, nom, val, disp, img) in cur.fetchall():
                imagen_url = _normalize_image_url(img)
                elementos.append({
                    'id': eid,
                    'nombre': nom,
                    'valor': float(val or 0),
                    'disponible': int(disp or 0),
                    'imagen': imagen_url
                })
        else:
            cur.execute("""
                SELECT ElementoId, NombreElemento, ValorUnitario, CantidadDisponible
                FROM dbo.ElementosPublicitarios
                WHERE Activo = 1 AND CantidadDisponible > 0
                ORDER BY NombreElemento
            """)
            for (eid, nom, val, disp) in cur.fetchall():
                elementos.append({
                    'id': eid,
                    'nombre': nom,
                    'valor': float(val or 0),
                    'disponible': int(disp or 0),
                    'imagen': None
                })
    except Exception as e:
        logger.error("Error cargando elementos: [error]")
        flash("Error interno al cargar elementos. Intenta de nuevo o contacta al administrador.", "danger")
    finally:
        try:
            if cur: cur.close()
            if conn: conn.close()
        except:
            pass

    return safe_render_template(
        'prestamos/crear.html',
        elementos=elementos,
        solicitante_id=solicitante_id,
        solicitante_nombre=solicitante_nombre,
        oficina_id=oficina_id,
        oficina_nombre=oficina_nombre,
        fecha_minima=fecha_minima
    )

# =========================
# Rutas de Aprobación/Rechazo/Devolución con Notificaciones
# =========================

@prestamos_bp.route('/<int:prestamo_id>/aprobar', methods=['POST'])
def aprobar_prestamo(prestamo_id):
    """Aprobar un préstamo pendiente"""
    if not _require_login():
        return jsonify({'success': False, 'message': 'No autorizado'}), 401
    
    if not can_access('prestamos', 'approve'):
        return jsonify({'success': False, 'message': 'No tienes permisos para aprobar préstamos'}), 403
    
    conn = cur = None
    try:
        conn = get_database_connection()
        cur = conn.cursor()
        
        # Obtener info del préstamo ANTES de aprobar
        prestamo_info = _obtener_info_prestamo_completa(prestamo_id)
        
        cur.execute("""
            SELECT pe.Estado, pe.ElementoId, pe.CantidadPrestada
            FROM dbo.PrestamosElementos pe
            WHERE pe.PrestamoId = ? AND pe.Activo = 1
        """, (prestamo_id,))
        
        row = cur.fetchone()
        if not row:
            return jsonify({'success': False, 'message': 'Préstamo no encontrado'}), 404
        
        estado_actual = row[0]
        elemento_id = row[1]
        cantidad_prestada = row[2]
        
       
        if estado_actual != 'PENDIENTE':
            return jsonify({'success': False, 'message': f'El préstamo ya está en estado: {estado_actual}'}), 400
        
        usuario_aprobador = session.get('usuario_nombre', 'Sistema')
        
        cur.execute("""
            UPDATE dbo.PrestamosElementos
            SET Estado = 'APROBADO',
                UsuarioAprobador = ?,
                FechaAprobacion = GETDATE(),
                ObservacionesAprobacion = ISNULL(ObservacionesAprobacion, '') + ' - Aprobado por: ' + ? + ' (' + CONVERT(VARCHAR, GETDATE(), 120) + ')'
            WHERE PrestamoId = ? AND Activo = 1
        """, (usuario_aprobador, usuario_aprobador, prestamo_id))
        
        conn.commit()
        
        # ====== NOTIFICACIÓN: Préstamo aprobado ======
        if NOTIFICACIONES_ACTIVAS and prestamo_info:
            try:
                ok = NotificationService.notificar_cambio_estado_prestamo(
                    prestamo_info,
                    'APROBADO',
                    usuario_aprobador
                )
                (logger.info("📧 Notificación OK: Préstamo #%s aprobado", sanitizar_log_text(prestamo_id)) if ok else logger.warning("📧 Notificación FAIL: Préstamo #%s aprobado", sanitizar_log_text(prestamo_id)))
            except Exception as e:
                logger.error("Error enviando notificación de aprobación préstamo: [error]")
        # =============================================
        
        return jsonify({
            'success': True,
            'message': 'Préstamo aprobado exitosamente',
            'prestamo_id': prestamo_id,
            'estado': 'APROBADO',
            'usuario_aprobador': usuario_aprobador,
            'fecha_aprobacion': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
        
    except Exception as e:
        try:
            if conn: conn.rollback()
        except:
            pass
        
        logger.info("❌ Error aprobando préstamo {prestamo_id}: [error]")
        return jsonify({'success': False, 'message': 'Error interno al aprobar el préstamo.'}), 500
    finally:
        try:
            if cur: cur.close()
            if conn: conn.close()
        except:
            pass

@prestamos_bp.route('/<int:prestamo_id>/aprobar_parcial', methods=['POST'])
def aprobar_parcial_prestamo(prestamo_id):
    """Aprobar parcialmente un préstamo pendiente"""
    if not _require_login():
        return jsonify({'success': False, 'message': 'No autorizado'}), 401
    
    if not can_access('prestamos', 'approve'):
        return jsonify({'success': False, 'message': 'No tienes permisos para aprobar préstamos'}), 403
    
    data = request.get_json()
    if not data or 'cantidad_aprobada' not in data:
        return jsonify({'success': False, 'message': 'Cantidad aprobada requerida'}), 400
    
    try:
        cantidad_aprobada = int(data['cantidad_aprobada'])
        if cantidad_aprobada <= 0:
            return jsonify({'success': False, 'message': 'La cantidad aprobada debe ser mayor a 0'}), 400
    except:
        return jsonify({'success': False, 'message': 'Cantidad inválida'}), 400
    
    conn = cur = None
    try:
        conn = get_database_connection()
        cur = conn.cursor()
        
        # Obtener info del préstamo ANTES de aprobar
        prestamo_info = _obtener_info_prestamo_completa(prestamo_id)
        
        cur.execute("""
            SELECT pe.Estado, pe.ElementoId, pe.CantidadPrestada, el.NombreElemento
            FROM dbo.PrestamosElementos pe
            INNER JOIN dbo.ElementosPublicitarios el ON pe.ElementoId = el.ElementoId
            WHERE pe.PrestamoId = ? AND pe.Activo = 1
        """, (prestamo_id,))
        
        row = cur.fetchone()
        if not row:
            return jsonify({'success': False, 'message': 'Préstamo no encontrado'}), 404
        
        estado_actual = row[0]
        elemento_id = row[1]
        cantidad_total = row[2]
        nombre_elemento = row[3]
        
        # Cambiado de 'PRESTADO' a 'PENDIENTE'
        if estado_actual != 'PENDIENTE':
            return jsonify({'success': False, 'message': f'El préstamo ya está en estado: {estado_actual}'}), 400
        
        if cantidad_aprobada > cantidad_total:
            return jsonify({
                'success': False, 
                'message': f'La cantidad aprobada ({cantidad_aprobada}) no puede exceder la cantidad solicitada ({cantidad_total})'
            }), 400
        
        usuario_aprobador = session.get('usuario_nombre', 'Sistema')
        
        cur.execute("""
            SELECT CantidadDisponible 
            FROM dbo.ElementosPublicitarios 
            WHERE ElementoId = ? AND Activo = 1
        """, (elemento_id,))
        
        stock_row = cur.fetchone()
        if not stock_row:
            return jsonify({'success': False, 'message': 'Elemento no encontrado'}), 404
        
        stock_disponible = stock_row[0] or 0
        
        if cantidad_aprobada > stock_disponible:
            return jsonify({
                'success': False, 
                'message': f'Stock insuficiente. Disponible: {stock_disponible}, Solicitado: {cantidad_aprobada}'
            }), 400
        
        cur.execute("""
            UPDATE dbo.PrestamosElementos
            SET Estado = 'APROBADO_PARCIAL',
                CantidadPrestada = ?,
                UsuarioAprobador = ?,
                FechaAprobacion = GETDATE(),
                ObservacionesAprobacion = ISNULL(ObservacionesAprobacion, '') + ' - Aprobado parcialmente por: ' + ? + ' (' + CONVERT(VARCHAR, GETDATE(), 120) + ') Cantidad: ' + CAST(? AS NVARCHAR) + ' de ' + CAST(? AS NVARCHAR),
                Observaciones = ISNULL(Observaciones, '') + ' - Aprobado parcialmente: ' + CAST(? AS NVARCHAR) + ' de ' + CAST(? AS NVARCHAR)
            WHERE PrestamoId = ? AND Activo = 1
        """, (cantidad_aprobada, usuario_aprobador, usuario_aprobador, 
              cantidad_aprobada, cantidad_total, cantidad_aprobada, cantidad_total, prestamo_id))
        
        cur.execute("""
            UPDATE dbo.ElementosPublicitarios
            SET CantidadDisponible = CantidadDisponible - ?
            WHERE ElementoId = ? AND Activo = 1
        """, (cantidad_aprobada, elemento_id))
        
        if cantidad_aprobada < cantidad_total:
            diferencia = cantidad_total - cantidad_aprobada
            cur.execute("""
                UPDATE dbo.ElementosPublicitarios
                SET CantidadDisponible = CantidadDisponible + ?
                WHERE ElementoId = ? AND Activo = 1
            """, (diferencia, elemento_id))
        
        conn.commit()
        
        # ====== NOTIFICACIÓN: Préstamo aprobado parcialmente ======
        if NOTIFICACIONES_ACTIVAS and prestamo_info:
            try:
                ok = NotificationService.notificar_cambio_estado_prestamo(
                    prestamo_info,
                    'APROBADO_PARCIAL',
                    usuario_aprobador,
                    f'Cantidad aprobada: {cantidad_aprobada} de {cantidad_total}'
                )
                (logger.info("📧 Notificación OK: Préstamo #%s aprobado parcialmente", sanitizar_log_text(prestamo_id)) if ok else logger.warning("📧 Notificación FAIL: Préstamo #%s aprobado parcialmente", sanitizar_log_text(prestamo_id)))
            except Exception as e:
                logger.error("Error enviando notificación de aprobación parcial préstamo: [error]")
        # =============================================
        
        return jsonify({
            'success': True,
            'message': f'Préstamo aprobado parcialmente ({cantidad_aprobada} de {cantidad_total})',
            'prestamo_id': prestamo_id,
            'estado': 'APROBADO_PARCIAL',
            'cantidad_aprobada': cantidad_aprobada,
            'cantidad_original': cantidad_total,
            'usuario_aprobador': usuario_aprobador,
            'fecha_aprobacion': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
        
    except Exception as e:
        try:
            if conn: conn.rollback()
        except:
            pass
        
        logger.info("❌ Error aprobando parcialmente préstamo {prestamo_id}: [error]")
        return jsonify({'success': False, 'message': 'Error interno al aprobar parcialmente.'}), 500
    finally:
        try:
            if cur: cur.close()
            if conn: conn.close()
        except:
            pass

@prestamos_bp.route('/<int:prestamo_id>/rechazar', methods=['POST'])
def rechazar_prestamo(prestamo_id):
    """Rechazar un préstamo pendiente"""
    if not _require_login():
        return jsonify({'success': False, 'message': 'No autorizado'}), 401
    
    if not can_access('prestamos', 'reject'):
        return jsonify({'success': False, 'message': 'No tienes permisos para rechazar préstamos'}), 403
    
    data = request.get_json()
    if not data or 'observacion' not in data:
        return jsonify({'success': False, 'message': 'Observación requerida'}), 400
    
    observacion = (data['observacion'] or '').strip()
    if not observacion:
        return jsonify({'success': False, 'message': 'La observación es obligatoria'}), 400
    
    conn = cur = None
    try:
        conn = get_database_connection()
        cur = conn.cursor()
        
        # Obtener info del préstamo ANTES de rechazar
        prestamo_info = _obtener_info_prestamo_completa(prestamo_id)
        
        cur.execute("""
            SELECT pe.Estado, pe.ElementoId, pe.CantidadPrestada
            FROM dbo.PrestamosElementos pe
            WHERE pe.PrestamoId = ? AND pe.Activo = 1
        """, (prestamo_id,))
        
        row = cur.fetchone()
        if not row:
            return jsonify({'success': False, 'message': 'Préstamo no encontrado'}), 404
        
        estado_actual = row[0]
        elemento_id = row[1]
        cantidad_prestada = row[2]
        
        # Cambiado de 'PRESTADO' a 'PENDIENTE'
        if estado_actual != 'PENDIENTE':
            return jsonify({'success': False, 'message': f'El préstamo ya está en estado: {estado_actual}'}), 400
        
        usuario_rechazador = session.get('usuario_nombre', 'Sistema')
        
        cur.execute("""
            UPDATE dbo.PrestamosElementos
            SET Estado = 'RECHAZADO',
                UsuarioRechazador = ?,
                FechaRechazo = GETDATE(),
                MotivoRechazo = ?,
                Observaciones = ISNULL(Observaciones, '') + ' - RECHAZADO por: ' + ? + ' (' + CONVERT(VARCHAR, GETDATE(), 120) + ') Motivo: ' + ?
            WHERE PrestamoId = ? AND Activo = 1
        """, (usuario_rechazador, observacion, usuario_rechazador, observacion, prestamo_id))
        
        # Devolver al inventario
        cur.execute("""
            UPDATE dbo.ElementosPublicitarios
            SET CantidadDisponible = CantidadDisponible + ?
            WHERE ElementoId = ? AND Activo = 1
        """, (cantidad_prestada, elemento_id))
        
        conn.commit()
        
        # ====== NOTIFICACIÓN: Préstamo rechazado ======
        if NOTIFICACIONES_ACTIVAS and prestamo_info:
            try:
                ok = NotificationService.notificar_cambio_estado_prestamo(
                    prestamo_info,
                    'RECHAZADO',
                    usuario_rechazador,
                    observacion
                )
                (logger.info("📧 Notificación OK: Préstamo #%s rechazado", sanitizar_log_text(prestamo_id)) if ok else logger.warning("📧 Notificación FAIL: Préstamo #%s rechazado", sanitizar_log_text(prestamo_id)))
            except Exception as e:
                logger.error("Error enviando notificación de rechazo préstamo: [error]")
        # =============================================
        
        return jsonify({
            'success': True,
            'message': 'Préstamo rechazado exitosamente',
            'prestamo_id': prestamo_id,
            'estado': 'RECHAZADO',
            'usuario_rechazador': usuario_rechazador,
            'fecha_rechazo': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'motivo_rechazo': observacion
        })
        
    except Exception as e:
        try:
            if conn: conn.rollback()
        except:
            pass
        
        logger.info("❌ Error rechazando préstamo {prestamo_id}: [error]")
        return jsonify({'success': False, 'message': 'Error interno al rechazar el préstamo.'}), 500
    finally:
        try:
            if cur: cur.close()
            if conn: conn.close()
        except:
            pass

@prestamos_bp.route('/<int:prestamo_id>/devolucion', methods=['POST'])
def registrar_devolucion_prestamo(prestamo_id):
    """Registrar devolución de un préstamo aprobado"""
    if not _require_login():
        return jsonify({'success': False, 'message': 'No autorizado'}), 401
    
    if not can_access('prestamos', 'return'):
        return jsonify({'success': False, 'message': 'No tienes permisos para registrar devoluciones'}), 403
    
    data = request.get_json()
    observacion = (data.get('observacion') or '').strip() if data else ''
    
    conn = cur = None
    try:
        conn = get_database_connection()
        cur = conn.cursor()
        
        # Obtener info del préstamo ANTES de devolver
        prestamo_info = _obtener_info_prestamo_completa(prestamo_id)
        
        cur.execute("""
            SELECT pe.Estado, pe.ElementoId, pe.CantidadPrestada
            FROM dbo.PrestamosElementos pe
            WHERE pe.PrestamoId = ? AND pe.Activo = 1
        """, (prestamo_id,))
        
        row = cur.fetchone()
        if not row:
            return jsonify({'success': False, 'message': 'Préstamo no encontrado'}), 404
        
        estado_actual = row[0]
        elemento_id = row[1]
        cantidad_prestada = row[2]
        
        if estado_actual not in ['APROBADO', 'APROBADO_PARCIAL']:
            return jsonify({
                'success': False, 
                'message': f'No se puede registrar devolución para préstamo en estado: {estado_actual}. Solo se puede devolver préstamos aprobados.'
            }), 400
        
        usuario_devolucion = session.get('usuario_nombre', 'Sistema')
        
        cur.execute("""
            UPDATE dbo.PrestamosElementos
            SET Estado = 'DEVUELTO',
                UsuarioDevolucion = ?,
                FechaDevolucionReal = GETDATE(),
                Observaciones = ISNULL(Observaciones, '') + CASE WHEN ? != '' THEN ' - DEVUELTO por: ' + ? + ' (' + CONVERT(VARCHAR, GETDATE(), 120) + ') Observación: ' + ? ELSE ' - DEVUELTO por: ' + ? + ' (' + CONVERT(VARCHAR, GETDATE(), 120) + ')' END
            WHERE PrestamoId = ? AND Activo = 1
        """, (usuario_devolucion, observacion, usuario_devolucion, observacion, usuario_devolucion, prestamo_id))
        
        # Devolver al inventario
        cur.execute("""
            UPDATE dbo.ElementosPublicitarios
            SET CantidadDisponible = CantidadDisponible + ?
            WHERE ElementoId = ? AND Activo = 1
        """, (cantidad_prestada, elemento_id))
        
        conn.commit()
        
        # ====== NOTIFICACIÓN: Devolución registrada ======
        if NOTIFICACIONES_ACTIVAS and prestamo_info:
            try:
                ok = NotificationService.notificar_cambio_estado_prestamo(
                    prestamo_info,
                    'DEVUELTO',
                    usuario_devolucion,
                    observacion if observacion else 'Devolución completada'
                )
                logger.info("📧 Notificación enviada: Devolución préstamo #%s", sanitizar_log_text(prestamo_id))
            except Exception as e:
                logger.error("Error enviando notificación de devolución préstamo: [error]")
        # =============================================
        
        return jsonify({
            'success': True,
            'message': 'Devolución registrada exitosamente',
            'prestamo_id': prestamo_id,
            'estado': 'DEVUELTO',
            'usuario_devolucion': usuario_devolucion,
            'fecha_devolucion': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'observacion': observacion if observacion else 'Sin observación'
        })
        
    except Exception as e:
        try:
            if conn: conn.rollback()
        except:
            pass
        
        logger.info("❌ Error registrando devolución préstamo {prestamo_id}: [error]")
        return jsonify({'success': False, 'message': 'Error interno al registrar la devolución.'}), 500
    finally:
        try:
            if cur: cur.close()
            if conn: conn.close()
        except:
            pass

# =========================
# Ruta: Ver detalles de préstamo
# =========================
@prestamos_bp.route('/<int:prestamo_id>', methods=['GET'])
def ver_prestamo(prestamo_id):
    """Ver detalles de un préstamo específico"""
    if not _require_login():
        return _redirect_login()
    
    detalle = _fetch_detalle(prestamo_id)
    if not detalle:
        flash('Préstamo no encontrado', 'danger')
        return redirect('/prestamos')
    
    return safe_render_template('prestamos/detalle.html', prestamo=detalle)

# =========================
# Ruta: Crear Material (AJAX + Tradicional)
# =========================
@prestamos_bp.route('/elementos/crearmaterial', methods=['GET', 'POST'])
def crear_material_prestamo():
    """Ruta para crear materiales en el módulo de préstamos"""
    if not _require_login():
        if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': 'No autorizado'}), 401
        return _redirect_login()
    
    if not can_access('materiales', 'create'):
        if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': '❌ No tienes permisos para crear materiales'}), 403
        flash('❌ No tienes permisos para crear materiales', 'danger')
        return redirect('/prestamos')

    # Restricción por oficinas específicas
    if _has_role('oficina_pereira', 'oficina_neiva', 'oficina_kennedy', 'oficina_bucaramanga', 
                 'oficina_polo_club', 'oficina_nogal', 'oficina_tunja', 'oficina_lourdes',
                 'oficina_cartagena', 'oficina_morato', 'oficina_medellin', 'oficina_cedritos'):
        if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': 'No tiene permisos para crear materiales publicitarios'}), 403
        flash('No tiene permisos para crear materiales publicitarios', 'danger')
        return redirect('/prestamos')
    
    if request.method == 'GET':
        return safe_render_template('prestamos/elemento_crear.html')

    # POST: Crear material
    nombre_elemento_raw = (request.form.get('nombre_elemento') or '').strip()
    nombre_elemento = nombre_elemento_raw.strip()

    valor_unitario_str = request.form.get('valor_unitario', '0')
    cantidad_disp_str = request.form.get('cantidad_disponible', '0')
    cantidad_minima_str = request.form.get('cantidad_minima', '0')
    imagen = request.files.get('imagen')

    # OFICINA FIJA: COQ (ID 1)
    oficina_id = 1
    usuario_nombre = (session.get('usuario_nombre') or 'administrador').strip() or 'administrador'

    # Validaciones
    try:
        valor_unitario = float(valor_unitario_str) if valor_unitario_str else 0.0
        cantidad_disp = int(cantidad_disp_str) if cantidad_disp_str else 0
        cantidad_minima = int(cantidad_minima_str) if cantidad_minima_str else 0
    except:
        msg = 'Valor unitario o cantidad no válidos.'
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': msg}), 400
        flash(msg, 'warning')
        return redirect('/prestamos/elementos/crearmaterial')

    if not nombre_elemento or valor_unitario <= 0 or cantidad_disp < 0 or cantidad_minima < 0:
        msg = 'Complete nombre, valor (>0) y stock (>=0).'
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': msg}), 400
        flash(msg, 'warning')
        return redirect('/prestamos/elementos/crearmaterial')

    # Guardar imagen
    ruta_imagen = None
    if imagen and imagen.filename:
        try:
            filename = secure_filename(imagen.filename)
            if not allowed_file(filename):
                msg = 'Formato de archivo no permitido. Use JPG, PNG, GIF o WEBP.'
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return jsonify({'success': False, 'message': msg}), 400
                flash(msg, 'warning')
                return redirect('/prestamos/elementos/crearmaterial')

            static_dir = current_app.static_folder
            upload_dir = os.path.join(static_dir, 'uploads', 'elementos')
            os.makedirs(upload_dir, exist_ok=True)

            file_path = os.path.join(upload_dir, filename)
            imagen.save(file_path)

            ruta_imagen = f'uploads/elementos/{filename}'
        except Exception as e:
            msg = 'Error interno al guardar la imagen'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'message': msg}), 500
            flash(msg, 'danger')
            return redirect('/prestamos/elementos/crearmaterial')

    conn = cur = None
    try:
        conn = get_database_connection()
        cur = conn.cursor()

        # ============================
        # VERIFICAR SI YA EXISTE EL MATERIAL EN ESTA OFICINA
        # ============================
        cur.execute("""
            SELECT COUNT(*) 
            FROM dbo.ElementosPublicitarios 
            WHERE NombreElemento = ? AND OficinaCreadoraId = ? AND Activo = 1
        """, (nombre_elemento, oficina_id))
        
        if cur.fetchone()[0] > 0:
            error_message = f'Ya existe un material con el nombre "{nombre_elemento}" en esta oficina'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'message': error_message}), 409
            flash(error_message, 'danger')
            return redirect('/prestamos/elementos/crearmaterial')

        # ============================
        # INSERTAR MATERIAL
        # ============================
        columnas = [
            "NombreElemento", "ValorUnitario", "CantidadDisponible", "CantidadMinima",
            "OficinaCreadoraId", "Activo", "FechaCreacion", "UsuarioCreador"
        ]
        valores = [
            nombre_elemento, valor_unitario, cantidad_disp, cantidad_minima,
            oficina_id, 1, datetime.now(), usuario_nombre
        ]

        if ruta_imagen:
            columnas.append("RutaImagen")
            valores.append(ruta_imagen)

        sql = f"""
            INSERT INTO dbo.ElementosPublicitarios
            ({", ".join(columnas)})
            VALUES ({", ".join(["?"] * len(columnas))})
        """

        cur.execute(sql, tuple(valores))
        conn.commit()

        # Calcular valor total para el mensaje
        valor_total = valor_unitario * cantidad_disp
        success_message = f'✅ Elemento "{nombre_elemento}" creado correctamente. Valor total: ${valor_total:.2f}'
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({
                'success': True,
                'message': success_message,
                'data': {'nombre': nombre_elemento, 'valor_total': valor_total}
            })

        flash(success_message, 'success')
        return redirect('/prestamos/elementos/crearmaterial')

    except Exception as e:
        try:
            if conn: conn.rollback()
        except:
            pass

        error_parts = []
        try:
            for _a in getattr(e, 'args', ()):
                if isinstance(_a, str):
                    error_parts.append(_a.lower())
        except Exception:
            error_parts = []
        error_text = ' '.join(error_parts) if error_parts else ''
        error_message = 'Error interno al crear material.'

        logger.error("Error en crear_material_prestamo: [error]")
        # Si es error de duplicado, mensaje más específico
        if ('duplicate' in error_text) or ('unique' in error_text) or ('2627' in error_text) or ('2601' in error_text):
            error_message = f'Ya existe un material con el nombre "{nombre_elemento}" en esta oficina'

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': error_message}), 500

        flash(error_message, 'danger')
        return redirect('/prestamos/elementos/crearmaterial')

    finally:
        try:
            if cur: cur.close()
            if conn: conn.close()
        except:
            pass

# =========================
# Exportaciones
# =========================
@prestamos_bp.route('/exportar/excel')
def exportar_prestamos_excel():
    """Exporta los préstamos filtrados a Excel.

    Filtros soportados (mismos que la vista):
    - estado, oficina, material, solicitante, fecha_inicio, fecha_fin
    """
    if not (can_access('prestamos', 'view') or can_access('prestamos', 'view_all') or can_access('prestamos', 'view_own')):
        flash('❌ No tienes permisos para exportar préstamos', 'danger')
        return redirect('/prestamos')

    try:
        filtro_estado = (request.args.get('estado') or '').strip()
        filtro_oficina = (request.args.get('oficina') or 'todas').strip() or 'todas'
        filtro_material = (request.args.get('material') or '').strip()
        filtro_solicitante = (request.args.get('solicitante') or '').strip()
        filtro_fecha_inicio = (request.args.get('fecha_inicio') or '').strip()
        filtro_fecha_fin = (request.args.get('fecha_fin') or '').strip()

        from utils.permissions import user_can_view_all
        can_all = user_can_view_all()

        oficina_id = None
        if can_all:
            if filtro_oficina and filtro_oficina != 'todas':
                try:
                    oficina_id = int(filtro_oficina)
                except Exception:
                    oficina_id = None
        else:
            oficina_id = session.get('oficina_id')

        prestamos = _fetch_prestamos(filtro_estado or None, oficina_id)

        prestamos = _apply_extra_filters(
            prestamos,
            filtro_material=filtro_material,
            filtro_solicitante=filtro_solicitante,
            fecha_inicio=filtro_fecha_inicio,
            fecha_fin=filtro_fecha_fin
        )

        if not prestamos:
            flash('No hay préstamos para exportar con los filtros seleccionados.', 'warning')
            return redirect('/prestamos' + (request.query_string.decode('utf-8', errors='ignore') and ('?' + request.query_string.decode('utf-8', errors='ignore')) or ''))

        # Preferir pandas si está disponible (mantiene compatibilidad con lo ya instalado)
        if HAS_PANDAS:
            columnas = [
                'ID', 'Material', 'Cantidad', 'Valor Unitario', 'Subtotal',
                'Solicitante', 'Oficina', 'Fecha Préstamo', 'Fecha Devolución Esperada', 'Estado',
                'Usuario Aprobador', 'Fecha Aprobación', 'Usuario Rechazador', 'Fecha Rechazo',
                'Usuario Devolución', 'Fecha Devolución Real'
            ]
            data = [{
                'ID': p.get('id', ''),
                'Material': p.get('material', ''),
                'Cantidad': p.get('cantidad', 0),
                'Valor Unitario': float(p.get('valor_unitario', 0)),
                'Subtotal': float(p.get('subtotal', 0)),
                'Solicitante': p.get('solicitante_nombre', ''),
                'Oficina': p.get('oficina_nombre', ''),
                'Fecha Préstamo': p.get('fecha', ''),
                'Fecha Devolución Esperada': p.get('fecha_prevista', ''),
                'Estado': p.get('estado', ''),
                'Usuario Aprobador': p.get('usuario_aprobador', ''),
                'Fecha Aprobación': p.get('fecha_aprobacion', ''),
                'Usuario Rechazador': p.get('usuario_rechazador', ''),
                'Fecha Rechazo': p.get('fecha_rechazo', ''),
                'Usuario Devolución': p.get('usuario_devolucion', ''),
                'Fecha Devolución Real': p.get('fecha_devolucion_real', '')
            } for p in prestamos]

            df = pd.DataFrame(data, columns=columnas)

            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Préstamos', index=False)
                ws = writer.sheets['Préstamos']
                for col_cells in ws.columns:
                    max_len = 0
                    col_letter = col_cells[0].column_letter
                    for c in col_cells:
                        try:
                            max_len = max(max_len, len(str(c.value)) if c.value is not None else 0)
                        except Exception:
                            pass
                    ws.column_dimensions[col_letter].width = max_len + 2

            output.seek(0)
        else:
            # Fallback sin pandas: openpyxl puro
            try:
                from openpyxl import Workbook
                from openpyxl.utils import get_column_letter
            except Exception:
                flash('Exportar a Excel requiere pandas u openpyxl.', 'warning')
                return redirect('/prestamos')

            wb = Workbook()
            ws = wb.active
            ws.title = 'Préstamos'

            headers = [
                'ID', 'Material', 'Cantidad', 'Valor Unitario', 'Subtotal',
                'Solicitante', 'Oficina', 'Fecha Préstamo', 'Fecha Devolución Esperada', 'Estado'
            ]
            ws.append(headers)

            for p in prestamos:
                ws.append([
                    p.get('id', ''),
                    p.get('material', ''),
                    p.get('cantidad', 0),
                    float(p.get('valor_unitario', 0)),
                    float(p.get('subtotal', 0)),
                    p.get('solicitante_nombre', ''),
                    p.get('oficina_nombre', ''),
                    str(p.get('fecha', '') or ''),
                    str(p.get('fecha_prevista', '') or ''),
                    p.get('estado', '')
                ])

            # Auto-ajuste simple de columnas
            for col_idx, _ in enumerate(headers, start=1):
                col_letter = get_column_letter(col_idx)
                max_len = 0
                for cell in ws[col_letter]:
                    try:
                        max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = max_len + 2

            output = BytesIO()
            wb.save(output)
            output.seek(0)

        fecha_actual = datetime.now().strftime('%Y-%m-%d')
        filename = f'prestamos_{fecha_actual}.xlsx'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        logger.info("❌ Error exportando préstamos a Excel: [error]")
        flash('Error al exportar el reporte de préstamos a Excel', 'danger')
        return redirect('/prestamos')

@prestamos_bp.route('/exportar/pdf')
def exportar_prestamos_pdf():
    """Exporta los préstamos filtrados a PDF (WeasyPrint).

    Filtros soportados (mismos que la vista):
    - estado, oficina, material, solicitante, fecha_inicio, fecha_fin
    """
    if not (can_access('prestamos', 'view') or can_access('prestamos', 'view_all') or can_access('prestamos', 'view_own')):
        flash('❌ No tienes permisos para exportar préstamos', 'danger')
        return redirect('/prestamos')

    try:
        if not HAS_WEASYPRINT:
            flash('Exportar a PDF requiere WeasyPrint instalado.', 'warning')
            return redirect('/prestamos')

        filtro_estado = (request.args.get('estado') or '').strip()
        filtro_oficina = (request.args.get('oficina') or 'todas').strip() or 'todas'
        filtro_material = (request.args.get('material') or '').strip()
        filtro_solicitante = (request.args.get('solicitante') or '').strip()
        filtro_fecha_inicio = (request.args.get('fecha_inicio') or '').strip()
        filtro_fecha_fin = (request.args.get('fecha_fin') or '').strip()

        from utils.permissions import user_can_view_all
        can_all = user_can_view_all()

        oficina_id = None
        if can_all:
            if filtro_oficina and filtro_oficina != 'todas':
                try:
                    oficina_id = int(filtro_oficina)
                except Exception:
                    oficina_id = None
        else:
            oficina_id = session.get('oficina_id')

        prestamos = _fetch_prestamos(filtro_estado or None, oficina_id)

        prestamos = _apply_extra_filters(
            prestamos,
            filtro_material=filtro_material,
            filtro_solicitante=filtro_solicitante,
            fecha_inicio=filtro_fecha_inicio,
            fecha_fin=filtro_fecha_fin
        )

        from html import escape as h

        filas_html = "\n".join(
            f"""<tr>
                <td>{h(str(p.get('id', '') or ''))}</td>
                <td>{h(str(p.get('material', '') or ''))}</td>
                <td>{h(str(p.get('cantidad', 0) or 0))}</td>
                <td>{h(str(p.get('solicitante_nombre', '') or ''))}</td>
                <td>{h(str(p.get('oficina_nombre', '') or ''))}</td>
                <td>{h(str(p.get('fecha', '') or ''))}</td>
                <td>{h(str(p.get('estado', '') or ''))}</td>
            </tr>"""
            for p in (prestamos or [])
        )

        html_content = f"""
        <html>
        <head>
            <meta charset="utf-8" />
            <style>
                body {{ font-family: Arial, sans-serif; font-size: 12px; }}
                h1 {{ margin: 0; }}
                .header {{ text-align: center; margin-bottom: 16px; }}
                table {{ width: 100%; border-collapse: collapse; }}
                th, td {{ border: 1px solid #ddd; padding: 6px; text-align: left; }}
                th {{ background-color: #f2f2f2; }}
                .meta {{ color: #555; font-size: 11px; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>Reporte de Préstamos</h1>
                <div class="meta">
                    Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}<br/>
                    Total de préstamos: {len(prestamos)}
                </div>
            </div>
            <table>
                <thead>
                    <tr>
                        <th>ID</th>
                        <th>Material</th>
                        <th>Cantidad</th>
                        <th>Solicitante</th>
                        <th>Oficina</th>
                        <th>Fecha</th>
                        <th>Estado</th>
                    </tr>
                </thead>
                <tbody>
                    {filas_html}
                </tbody>
            </table>
        </body>
        </html>
        """

        # Crear PDF temporal
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
        tmp_path = tmp.name
        tmp.close()
        HTML(string=html_content).write_pdf(tmp_path)

        @after_this_request
        def _remove_file(response):
            try:
                os.remove(tmp_path)
            except Exception:
                pass
            return response

        fecha_actual = datetime.now().strftime('%Y-%m-%d')
        filename = f'prestamos_{fecha_actual}.pdf'

        return send_file(
            tmp_path,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        logger.info("❌ Error exportando préstamos a PDF: [error]")
        flash('Error al exportar el reporte de préstamos a PDF', 'danger')
        return redirect('/prestamos')

# =========================
# API auxiliar: datos de un elemento
# =========================
@prestamos_bp.route('/api/elemento/<int:elemento_id>')
def api_elemento_info(elemento_id: int):
    """API para obtener información de un elemento publicitario"""
    if not _require_login():
        return jsonify({'ok': False, 'error': 'No autorizado'}), 401
    
    conn = cur = None
    try:
        conn = get_database_connection()
        cur = conn.cursor()
        
        img_col = _detect_image_column(cur)
        
        if img_col:
            cur.execute(f"""
                SELECT ElementoId, NombreElemento, ValorUnitario, CantidadDisponible, {img_col}
                FROM dbo.ElementosPublicitarios
                WHERE ElementoId = ? AND Activo = 1
            """, (elemento_id,))
        else:
            cur.execute("""
                SELECT ElementoId, NombreElemento, ValorUnitario, CantidadDisponible
                FROM dbo.ElementosPublicitarios
                WHERE ElementoId = ? AND Activo = 1
            """, (elemento_id,))
        
        row = cur.fetchone()
        if row:
            imagen_url = ""
            if img_col and len(row) >= 5:
                imagen_url = _normalize_image_url(row[4])
            
            return jsonify({
                'ok': True,
                'id': row[0],
                'nombre': row[1],
                'valor_unitario': float(row[2] or 0),
                'disponible': int(row[3] or 0),
                'imagen': imagen_url
            })
        else:
            return jsonify({'ok': False, 'error': 'Elemento no encontrado'}), 404
            
    except Exception as e:
        logger.info("Error en api_elemento_info: [error]")
        return jsonify({'ok': False, 'error': 'Error interno'}), 500
    finally:
        try:
            if cur: cur.close()
            if conn: conn.close()
        except: 
            pass

@prestamos_bp.route('/crearmaterial', methods=['GET'], endpoint='crearmaterial_legacy')
def crearmaterial_legacy():
    """Alias legacy (GET) para creación de material.

    Importante:
    - NO debe usar el endpoint 'crear_material' para evitar colisiones con el alias
      que apunta a la ruta canónica '/elementos/crearmaterial'.
    - Redirige a la ruta canónica para evitar lógica duplicada.
    """
    if not _require_login():
        return _redirect_login()

    if not can_access('prestamos', 'manage_materials'):
        flash('No tienes permisos para crear materiales', 'danger')
        return redirect('/prestamos')

    return redirect('/prestamos/elementos/crearmaterial')
# =========================
# Aliases de endpoints (compatibilidad con templates/base)
# =========================
def _registrar_aliases_endpoints():
    """Registra endpoints alternos sin afectar el dispatch.

    Motivo: En algunos templates/base se usan nombres cortos (p.ej. 'prestamos.crear')
    mientras que en el blueprint se definieron nombres largos (p.ej. 'prestamos.crear_prestamo').
    Para evitar BuildError, añadimos aliases con el MISMO path pero OTRO endpoint.
    """
    def _alias(rule: str, endpoint: str, view_func, methods=None):
        """Registra un alias sólo si el endpoint no existe.

        Seguridad/estabilidad:
        - Evita AssertionError por colisiones de endpoints (p.ej. en hot-reload o cambios de plantilla).
        - No sobreescribe endpoints existentes.
        """
        try:
            if endpoint in getattr(prestamos_bp, 'view_functions', {}):
                return
            if methods:
                prestamos_bp.add_url_rule(rule, endpoint=endpoint, view_func=view_func, methods=methods)
            else:
                prestamos_bp.add_url_rule(rule, endpoint=endpoint, view_func=view_func)
        except Exception:
            # No romper la carga del módulo por alias no crítico
            return

    # Listado
    _alias('/', 'index', listar_prestamos)
    _alias('/', 'listar', listar_prestamos)

    # Crear préstamo
    _alias('/crear', 'crear', crear_prestamo, methods=['GET', 'POST'])
    _alias('/crear', 'nuevo', crear_prestamo, methods=['GET', 'POST'])

    # Crear material (ruta canónica)
    _alias('/elementos/crearmaterial', 'crearmaterial', crear_material_prestamo, methods=['GET', 'POST'])
    _alias('/elementos/crearmaterial', 'crear_material', crear_material_prestamo, methods=['GET', 'POST'])

    # Export
    _alias('/exportar/excel', 'excel', exportar_prestamos_excel, methods=['GET'])
    _alias('/exportar/pdf', 'pdf', exportar_prestamos_pdf, methods=['GET'])

    # Detalle
    _alias('/<int:prestamo_id>', 'detalle', ver_prestamo, methods=['GET'])
    _alias('/<int:prestamo_id>', 'ver', ver_prestamo, methods=['GET'])

    # Acciones
    _alias('/<int:prestamo_id>/aprobar', 'aprobar', aprobar_prestamo, methods=['POST'])
    _alias('/<int:prestamo_id>/aprobar_parcial', 'aprobar_parcial', aprobar_parcial_prestamo, methods=['POST'])
    _alias('/<int:prestamo_id>/rechazar', 'rechazar', rechazar_prestamo, methods=['POST'])
    _alias('/<int:prestamo_id>/devolucion', 'devolver', registrar_devolucion_prestamo, methods=['POST'])
    _alias('/<int:prestamo_id>/devolver', 'devolver_legacy', registrar_devolucion_prestamo, methods=['POST'])

    # API elemento
    _alias('/api/elemento/<int:elemento_id>', 'elemento_info', api_elemento_info, methods=['GET'])

_registrar_aliases_endpoints()
